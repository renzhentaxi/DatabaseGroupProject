import sys
from openpyxl import *


def load_worksheet(fileName):
    print("loading worksheet from", fileName)
    return load_workbook(fileName).active


def tuple2csv(tuple):
    return "\t".join(map(str, tuple))


def locate_column(ws, columnName):
    columnName = columnName.lower()
    for i in range(1, ws.max_column + 1):
        if columnName == ws.cell(row=1, column=i).value.lower():
            return i
    print("Can not locate column named", columnName)
    return -1


def create_csv(path, data, fileName):
    print("creating csv file for", fileName)
    os.makedirs(os.path.dirname(path),exist_ok=True)

    fullName = path + fileName + ".csv"
    with open(fullName,"w+") as f:
        for row in data:
            f.write(tuple2csv(row) + "\n")
    print("created", len(data), "entries of",fileName)


# format should be first name, last name, user name, position
def load_employees(ws):
    print("loading employees")
    employees = []
    salesPersonCol = locate_column(ws, "Salesperson")
    userNameCol = locate_column(ws, "Salesperson Username")

    for i in range(2, ws.max_row + 1):
        salesPerson = ws.cell(row=i, column=salesPersonCol).value.partition(" ")
        userName = ws.cell(row=i, column=userNameCol).value
        firstName = salesPerson[0]
        lastName = salesPerson[2]

        emp = [firstName, lastName, userName, "rep"]
        if emp not in employees:
            employees.append(emp)
    for i in range (0, len(employees)):
        employees[i].insert(0, i)
    return employees


# format should be userName password
def generate_accounts(employees):
    print("Generating accounts")
    accounts = []
    for emp in employees:
        accountId = emp[0]
        accountName = emp[3]
        accountPassword = 1234
        account = [accountId, accountName, accountPassword]
        accounts.append(account)
    return accounts


def load_salesLeads(ws, accounts):
    print("loading salesleads")
    salesLeads = []

    firstNameCol = locate_column(ws, "first_name")
    lastNameCol = locate_column(ws, "last_name")
    comNameCol = locate_column(ws, "company_name")
    addressCol = locate_column(ws, "address")
    cityCol = locate_column(ws, "city")
    countyCol = locate_column(ws, "county")
    stateCol = locate_column(ws, "state")
    zipCol = locate_column(ws, "zip")
    phone1Col = locate_column(ws, "phone1")
    phone2Col = locate_column(ws, "phone2")
    emailCol = locate_column(ws, "email")
    webCol = locate_column(ws, "web")

    userNameCol = locate_column(ws, "salesperson username")
    userNameIdMap = dict(map(lambda x: (x[1],x[0]), accounts))

    salesId = 0

    for i in range(2, ws.max_row+1):
        firstName = ws.cell(row=i, column=firstNameCol).value
        lastName = ws.cell(row=i, column=lastNameCol).value
        comName = ws.cell(row=i, column=comNameCol).value
        address = ws.cell(row=i, column=addressCol).value
        city = ws.cell(row=i, column=cityCol).value
        county = ws.cell(row=i, column=countyCol).value
        state = ws.cell(row=i, column=stateCol).value
        zip = int(ws.cell(row=i, column=zipCol).value)
        phone1 = ws.cell(row=i, column=phone1Col).value
        phone2 = ws.cell(row=i, column=phone2Col).value
        email = ws.cell(row=i, column=emailCol).value
        web = ws.cell(row=i, column=webCol).value

        repUserName = ws.cell(row=i, column = userNameCol).value
        repId = userNameIdMap[repUserName]
        lead = (salesId, firstName, lastName, comName, address, city, county, state, zip, phone1, phone2, email, web, repId)

        salesLeads.append(lead)
        salesId +=1

    return salesLeads



argv = sys.argv
fileName = "./data/excels/rawData.xlsx"
csvPath = "./data/csv/"

if len(argv) > 3:
    print("Wrong arguments")
    print("Use CreateCSVScript excelLocation cvsLocation")
    raise Exception("Bad arguments")

if len(argv) >= 2:
    fileName = argv[1]
    if not os.path.isfile("./" + argv[1]):
        print("File: ", argv[1], "does not exist")
    if len(argv) == 3:
        csvPath = argv[2]


worksheet = load_worksheet(fileName)


employees = load_employees(worksheet)
accounts = generate_accounts(employees)
salesLeads = load_salesLeads(worksheet, accounts)

create_csv(csvPath,employees, "employees")
create_csv(csvPath,accounts, "accounts")
create_csv(csvPath,salesLeads, "salesleads")

print("done")
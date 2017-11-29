from openpyxl import *


def copyTo(wsFrom, fromRange, wsTo, toRange):
    for row in wsFrom[fromRange]:
        for cell in row:
            ws2.cell(row=cell.row,column=cell.col_idx,value = cell.value)

def ExtractCustomer(wbook):
    wb = Workbook()
    ws = wb.active
    copyTo(wbook.active, "a1:l50001", ws, 'a1:l50001')
    
def ExtractSalesRep(wbook):
    wb = Workbook()
    ws = wb.active
    entry =[]
    for row in wbook.active["m1:n50001"]:
        e = (row[0].value, row[1].value)
        
        if e not in entry:
            entry.append(e)
            ws.append(e)
    wb.save("rep.xlsx")

wb = load_workbook('rep.xlsx')
newwb = Workbook()
for cell in wb.active["a"]:
    newwb.active.cell(row = cell.row, column = cell.col_idx, value = cell.value)
newwb.active["b1"] = "Password"
newwb.save("Accounts.xlsx")
print("done")
